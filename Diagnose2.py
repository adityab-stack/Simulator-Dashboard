import openpyxl
wb = openpyxl.load_workbook("Automation_Data.xlsx", data_only=True)
ws = wb["Inv Data 2"]
print("dimensions:", ws.dimensions)
none_count = 0
first_real_row = None
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if row[0] is None:
        none_count += 1
        continue
    first_real_row = i
    print(f"First real data at row {i}:", row[:8])
    break
print(f"Blank rows before first real data: {none_count}")

# also show last few rows to confirm data exists somewhere
print("\nLast 3 rows of the sheet:")
last_row = ws.max_row
for r in range(last_row-2, last_row+1):
    row = [ws.cell(r, c).value for c in range(1, 9)]
    print(r, row)