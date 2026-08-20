"""
V1.2 data pipeline for the Snitch Control Room dashboard.

THREE ways to run this:

  A) Fully automatic, authenticated, IMPORTRANGE-safe (recommended -- set up once):
       python build_data.py
     Requires a Google Service Account (one-time ~10 min setup, see README).
     Save its JSON key as ./service_account.json next to this script, and share
     the Google Sheet with the service account's email (Viewer access).
     This goes through Google's live Sheets API, which correctly resolves
     IMPORTRANGE (and any other cross-sheet formula) because it's a real
     authenticated read, unlike a plain anonymous URL fetch.

  B) Automatic but UNSAFE for IMPORTRANGE-driven tabs:
       python build_data.py --export-url
     Downloads via the plain https://.../export?format=xlsx URL. Fine for
     Sales Data and Pipeline (plain formulas), but Inv Data 2 (IMPORTRANGE)
     will come back BLANK -- Google's export endpoint can't resolve
     IMPORTRANGE without an authenticated session. Kept only as a fallback.

  C) Offline / manual:
     python build_data.py --local
     Reads ./Automation_Data.xlsx if you've placed one there yourself.

Either way it writes data/data.json, which is the only thing index.html reads.

Setup for mode A (one-time):
  pip install google-api-python-client google-auth --break-system-packages
  1. console.cloud.google.com -> new/existing project -> enable "Google Sheets API"
  2. IAM & Admin -> Service Accounts -> Create Service Account (any name, no roles needed)
  3. That service account -> Keys -> Add Key -> Create new key -> JSON -> downloads a file
  4. Save that file as service_account.json in this same folder
  5. Open the file, copy the "client_email" value
  6. Share the actual Google Sheet with that email, Viewer access
  Done -- python build_data.py now works with zero manual downloads, forever.
"""
import json, datetime, sys, os, csv
import openpyxl

SHEET_ID = "1PqtpL9w2Tneon_-6zz7BGiW5YUz4fhDCrgn687r_CZw"
HERE = os.path.dirname(os.path.abspath(__file__))
LOCAL_XLSX = os.path.join(HERE, "Automation_Data.xlsx")
SERVICE_ACCOUNT_FILE = os.path.join(HERE, "service_account.json")
TARGETS_DIR = os.path.join(HERE, "data", "targets")
OUT = os.path.join(HERE, "data", "data.json")
TAB_NAMES = ["Sales Data", "Inv Data 2", "Pipeline"]

# Exact filenames expected in data/targets/ -- one per category family, each with
# a "Sheet1" tab at store/branch grain with monthly qty columns (AUG_2026.. JUN_2027).
TARGET_FILES = [
    "Jeans Planned Qty only.xlsx",
    "Shirts Planned Qty Only.xlsx",
    "Trousers Planned Qty only.xlsx",
    "tshirts Planned Qty Only.xlsx",
]
MONTH_COL_MAP = {
    "AUG_2026": "2026-08", "SEP_2026": "2026-09", "OCT_2026": "2026-10",
    "NOV_2026": "2026-11", "DEC_2026": "2026-12", "JAN_2027": "2027-01",
    "FEB_2027": "2027-02", "MAR_2027": "2027-03", "APR_2027": "2027-04",
    "MAY_2027": "2027-05", "JUN_2027": "2027-06",
}
# normalize target's channel labels to match Sales' channel naming where they're
# clearly the same thing; Warehouse kept as its own distinct channel since Sales
# doesn't have an equivalent yet. MP-SOR merged into Marketplace per instruction.
TARGET_CHANNEL_MAP = {"Online - Shopify": "Shopify"}
# NOTE: MP-SOR deliberately kept distinct from Marketplace here (not merged),
# per instruction -- the channel-mix split treats them as separate percentages.
# The live dashboard's "Marketplace" filter still sums MP + MP-SOR together at
# query time (same as it already does for Sales), so nothing visible changes.

NEW_TOTAL_FILE = os.path.join(HERE, "data", "targets_new_total.xlsx")


# ---------------------------------------------------------------------------
# Mode A: Google Sheets API via service account (authenticated, IMPORTRANGE-safe)
# ---------------------------------------------------------------------------
GSHEET_EPOCH = datetime.datetime(1899, 12, 30)

def _gsheet_serial_to_date(value):
    """Sheets API returns dates as serial-day floats (same epoch as Excel).
    Converts back to a python datetime so downstream code (month_key etc.)
    doesn't need to know the difference between this and an openpyxl cell."""
    if isinstance(value, (int, float)):
        return GSHEET_EPOCH + datetime.timedelta(days=value)
    return value


class SimpleSheet:
    """Minimal stand-in for an openpyxl worksheet, backed by raw API rows,
    so the same parsing code in build() works for either data source."""
    def __init__(self, rows, width, date_cols=()):
        # pad every row out to `width` columns (Sheets API omits trailing blanks)
        self.rows = []
        for row in rows:
            padded = list(row) + [None] * (width - len(row))
            for c in date_cols:
                if isinstance(padded[c], str) and padded[c] == "":
                    padded[c] = None
                elif isinstance(padded[c], (int, float)):
                    padded[c] = _gsheet_serial_to_date(padded[c])
            self.rows.append(tuple(padded))

    def iter_rows(self, min_row=1, max_row=None, values_only=True):
        start = min_row - 1
        end = max_row if max_row else len(self.rows)
        for r in self.rows[start:end]:
            yield r

    @property
    def max_row(self):
        return len(self.rows)


class SimpleWorkbook:
    def __init__(self, sheets_dict):
        self._sheets = sheets_dict
        self.sheetnames = list(sheets_dict.keys())

    def __getitem__(self, name):
        return self._sheets[name]


def fetch_via_sheets_api(sheet_id):
    from google.oauth2 import service_account
    from googleapiclient.discovery import build

    if not os.path.exists(SERVICE_ACCOUNT_FILE):
        print(f"ERROR: {SERVICE_ACCOUNT_FILE} not found.")
        print("See the module docstring (top of this file) for one-time setup steps.")
        sys.exit(1)

    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE,
        scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"]
    )
    service = build("sheets", "v4", credentials=creds)
    values_api = service.spreadsheets().values()

    # date columns (0-indexed) per tab, so serials get converted correctly
    date_cols = {"Sales Data": (0,), "Inv Data 2": (0,), "Pipeline": (0, 1, 20, 21)}
    widths = {"Sales Data": 15, "Inv Data 2": 17, "Pipeline": 28}

    sheets = {}
    for tab in TAB_NAMES:
        print(f"Fetching '{tab}' via Sheets API...")
        result = values_api.get(
            spreadsheetId=sheet_id, range=tab,
            valueRenderOption="UNFORMATTED_VALUE"
        ).execute()
        rows = result.get("values", [])
        sheets[tab] = SimpleSheet(rows, widths[tab], date_cols.get(tab, ()))
        print(f"  -> {len(rows)} rows (incl. header)")

    return SimpleWorkbook(sheets)


def download_from_gsheet(sheet_id, dest):
    import requests
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    print("Downloading workbook from Google Sheets (unauthenticated export)...")
    print("WARNING: this will NOT resolve IMPORTRANGE-driven tabs (e.g. Inv Data 2).")
    r = requests.get(url, timeout=60)
    r.raise_for_status()
    with open(dest, "wb") as f:
        f.write(r.content)
    print(f"Saved -> {dest}")


def month_key(dt):
    if dt is None:
        return None
    if isinstance(dt, str):
        return None  # e.g. "No Date mentioned" -- treat as undated, not a real month
    if hasattr(dt, "year") and dt.year < 2020:
        return None  # placeholder/epoch-error dates (e.g. 1970-01-01) -- not real
    return dt.strftime("%Y-%m")


def build_returns():
    """Loads the frozen, one-time 'learning' dataset (data/returns_learned.json,
    monthly Sales Qty / Return Qty at Channel x L1 x Category grain, 2026 only,
    MP-SOR already merged into Marketplace). This is NOT re-derived from a live
    source -- it's a static snapshot used as the starting point for the editable
    Returns tab in the dashboard, which is where the actually-applied rates
    (with caps, and any manual overrides) live from here on."""
    path = os.path.join(HERE, "data", "returns_learned.json")
    if not os.path.exists(path):
        print(f"NOTE: {path} not found -- Returns tab will start empty (0% assumed everywhere).")
        return []
    with open(path) as f:
        learned = json.load(f)
    print(f"Returns: loaded {len(learned)} frozen monthly learning rows.")
    return learned


def build_channel_mix_pct():
    """Computes OLD channel-mix % per (L1, Category, Month) from the 4
    Planned_Qty_only files -- summed across Meta/ASP-Bin/store, kept distinct
    per Channel (Warehouse excluded, not a sales channel)."""
    old_qty = {}  # (l1, cat, channel, month) -> qty
    any_found = False
    for fname in TARGET_FILES:
        path = os.path.join(TARGETS_DIR, fname)
        if not os.path.exists(path):
            print(f"NOTE: target file not found (skipping): data/targets/{fname}")
            continue
        any_found = True
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb["Sheet1"]
        col_idx = None
        for row in ws.iter_rows(values_only=True):
            if col_idx is None:
                if row and "L1_CATEGORY" in row:
                    col_idx = {name: i for i, name in enumerate(row)}
                continue
            l1 = row[col_idx["L1_CATEGORY"]]
            if l1 is None or l1 == "L1_CATEGORY":
                continue
            cat = (row[col_idx["CATEGORY"]] or "").strip().lower()
            channel_raw = row[col_idx["type"]]
            if channel_raw is None or channel_raw == "type" or channel_raw == "Warehouse":
                continue
            channel = TARGET_CHANNEL_MAP.get(channel_raw, channel_raw)
            for col_name, month_iso in MONTH_COL_MAP.items():
                idx = col_idx.get(col_name)
                if idx is None:
                    continue
                val = row[idx] or 0
                key = (l1, cat, channel, month_iso)
                old_qty[key] = old_qty.get(key, 0) + val
        wb.close()
    if not any_found:
        print("NOTE: no target files found in data/targets/ -- cannot compute channel mix.")
        return {}

    # totals per (l1, cat, month) across all channels
    totals = {}
    for (l1, cat, channel, month), qty in old_qty.items():
        tkey = (l1, cat, month)
        totals[tkey] = totals.get(tkey, 0) + qty

    pct = {}
    for (l1, cat, channel, month), qty in old_qty.items():
        total = totals.get((l1, cat, month), 0)
        if total > 0:
            pct[(l1, cat, channel, month)] = qty / total
    return pct


def build_new_totals():
    """Reads data/targets_new_total.xlsx (the new all-channel-combined Planned
    Qty numbers), aggregating away Meta1/2/3 and ASP Bin (ASP is parked for now)
    down to L1 x Category x Month."""
    if not os.path.exists(NEW_TOTAL_FILE):
        print("NOTE: data/targets_new_total.xlsx not found -- no new target totals to apply.")
        return {}
    wb = openpyxl.load_workbook(NEW_TOTAL_FILE, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    col_idx = None
    NEW_MONTH_COLS = {
        "Planned Qty AUG'26": "2026-08", "Planned Qty SEP'26": "2026-09",
        "Planned Qty OCT'26": "2026-10", "Planned Qty NOV'26": "2026-11",
        "Planned Qty DEC'26": "2026-12",
    }
    totals = {}
    for row in ws.iter_rows(values_only=True):
        if col_idx is None:
            if row and "Category" in row and "L1" in row:
                col_idx = {name: i for i, name in enumerate(row)}
            continue
        cat_raw = row[col_idx["Category"]]
        if cat_raw is None:
            continue
        cat = str(cat_raw).strip().lower()
        l1 = row[col_idx["L1"]]
        for col_name, month_iso in NEW_MONTH_COLS.items():
            idx = col_idx.get(col_name)
            if idx is None:
                continue
            val = row[idx] or 0
            key = (l1, cat, month_iso)
            totals[key] = totals.get(key, 0) + val
    return totals


def build_targets():
    """New target logic: preserves the OLD channel-mix % (from the 4
    Planned_Qty_only files) and applies it to the NEW total (from
    data/targets_new_total.xlsx), per L1 x Category x Month. The old files
    are now used only to derive the split ratio, not as target values
    themselves -- the new file's totals are authoritative."""
    pct = build_channel_mix_pct()
    new_totals = build_new_totals()

    targets = []
    for (l1, cat, month), new_total in new_totals.items():
        matching_channels = [k for k in pct if k[0] == l1 and k[1] == cat and k[3] == month]
        if not matching_channels:
            continue  # no old channel-mix data for this L1+Cat+Month -- can't split honestly
        for key in matching_channels:
            _, _, channel, _ = key
            share = pct[key]
            targets.append({"channel": channel, "l1": l1, "cat": cat, "month": month, "qty": new_total * share})
    return targets


def build(xlsx_path_or_workbook):
    if isinstance(xlsx_path_or_workbook, str):
        wb = openpyxl.load_workbook(xlsx_path_or_workbook, data_only=True)
    else:
        wb = xlsx_path_or_workbook  # already a loaded workbook (or SimpleWorkbook from the API path)

    # ---- Sales ----
    ws = wb["Sales Data"]
    sales = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        month, channel, l1, cat, m1, m2, m3, cogs, mrp, gross_sales, mrp_value, qty, cogs_sold = row[:13]
        sales.append({
            "month": month_key(month), "channel": channel, "l1": l1, "cat": cat,
            "m1": m1, "m2": m2, "m3": m3,
            "gross_sales": gross_sales or 0, "mrp_value": mrp_value or 0,
            "qty": qty or 0, "cogs_sold": cogs_sold or 0
        })

    # ---- Inventory ----
    # Kept at DAILY grain (not summed to month) because Closing(month) = the
    # latest dated snapshot within that month, per leaf combo -- summing days
    # together would wildly overstate stock. The frontend does the "latest
    # snapshot in month" pick, exactly matching the Overall tab's
    # SUMIFS(...,'CI v2'!date, MAXIFS(...)) formula.
    ws = wb["Inv Data 2"]
    inventory = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        inv_date, l1, cat, m1, m2, m3, mp_qty, online_qty, offline_qty, total_qty, mp_val, online_val, offline_val, total_val = row[:14]
        if l1 is None or cat is None:
            # stray/blank rows found in the source sheet (no L1 or Category) -- not real
            # inventory, skip so they don't pollute L1/Category rollups.
            continue
        d = inv_date.strftime("%Y-%m-%d") if hasattr(inv_date, "strftime") else str(inv_date)[:10]
        inventory.append({
            "date": d, "l1": l1, "cat": cat, "m1": m1, "m2": m2, "m3": m3,
            "mp_qty": mp_qty or 0, "online_qty": online_qty or 0, "offline_qty": offline_qty or 0,
            "total_qty": total_qty or 0,
            "mp_value": mp_val or 0, "online_value": online_val or 0, "offline_value": offline_val or 0,
            "total_value": total_val or 0
        })

    # ---- Pipeline (for Inwards) ----
    # Matched at the same leaf grain as inventory (L1+Cat+Meta1+Meta2+Meta3), keyed
    # by FDD_MONTH -- the month a batch is expected to land. No STATUS filter,
    # matching the original Overall-tab formula exactly.
    ws = wb["Pipeline"]
    pipeline = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        l1, cat, m1, m2, m3 = row[9], row[10], row[11], row[12], row[13]
        fdd_month, qty, cogs_unit = row[21], row[22], row[23]
        if l1 is None or cat is None:
            continue
        qty = qty or 0
        pipeline.append({
            "l1": l1, "cat": cat, "m1": m1, "m2": m2, "m3": m3,
            "fdd_month": month_key(fdd_month),  # None if undated
            "qty": qty,
            "cogs_value": qty * (cogs_unit or 0)
        })

    targets = build_targets()
    returns = build_returns()

    og_path = os.path.join(HERE, "data", "targets_og_fallback.json")
    targets_og = []
    if os.path.exists(og_path):
        with open(og_path) as f:
            targets_og = json.load(f)
    else:
        print("NOTE: data/targets_og_fallback.json not found -- no fallback for uncovered categories.")

    data = {
        "generated_at": datetime.datetime.now().isoformat(),
        "sales": sales,
        "inventory": inventory,
        "pipeline": pipeline,
        "targets": targets,        # primary source: the 4 category files, multi-channel
        "targets_og": targets_og,  # fallback: old Targets tab, Shopify-only, all categories
        "returns": returns,        # frozen monthly learning rows: Channel x L1 x Cat x Month (2026), sales_qty/return_qty
    }
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w") as f:
        json.dump(data, f)
    print(f"sales rows: {len(sales)}, inventory rows: {len(inventory)}, pipeline rows: {len(pipeline)}, "
          f"target rows (new files): {len(targets)}, target rows (OG fallback): {len(targets_og)}")
    if len(sales) > 0 and len(inventory) == 0:
        print("\n⚠️  WARNING: Sales has rows but Inventory is empty. This is the exact symptom")
        print("   of an unauthenticated download failing to resolve IMPORTRANGE on Inv Data 2.")
        print("   Fix: re-download Automation_Data.xlsx manually via your browser (File > Download")
        print("   > Microsoft Excel), then re-run with --local. See the module docstring for detail.\n")
    print(f"Wrote -> {OUT}")


if __name__ == "__main__":
    if "--local" in sys.argv:
        if not os.path.exists(LOCAL_XLSX):
            print(f"ERROR: {LOCAL_XLSX} not found.")
            sys.exit(1)
        build(LOCAL_XLSX)
    elif "--export-url" in sys.argv:
        download_from_gsheet(SHEET_ID, LOCAL_XLSX)
        build(LOCAL_XLSX)
    else:
        wb = fetch_via_sheets_api(SHEET_ID)
        build(wb)